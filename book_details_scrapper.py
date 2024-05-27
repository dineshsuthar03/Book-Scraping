import requests
from bs4 import BeautifulSoup
import csv
import logging
import traceback
from openpyxl import Workbook, load_workbook
import os

# Set up logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

# Create console handler and set level to INFO
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(formatter)
logger.addHandler(ch)

# Create file handler and set level to INFO
fh = logging.FileHandler('scraping.log')
fh.setLevel(logging.INFO)
fh.setFormatter(formatter)
logger.addHandler(fh)

def read_isbn_list_from_csv(filename):
    isbn_list = []
    with open(filename, mode='r') as file:
        reader = csv.reader(file)
        next(reader)  # Skip header row
        for row in reader:
            isbn = row[0].strip()  # Assuming ISBN is in the first column
            isbn_list.append(isbn)
    return isbn_list

def scrape_book_details(url, isbn):
    try:
        # Send a GET request to the URL
        logger.info(f"Fetching URL: {url}")
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for bad status codes
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching URL: {e}")
        logger.error(traceback.format_exc())  # Log traceback
        return None
    
    # Parse the HTML content
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Extract book details
    try:
        book_details = extract_book_details(soup)
    except Exception as e:
        logger.error(f"Error extracting book details: {e}")
        logger.error(traceback.format_exc())  # Log traceback
        return None
    
    if book_details:
        title = book_details.get('Title')
        if title:
            book_details['ISBN'] = isbn
            logger.info(f"New book found: {title} - ISBN: {isbn}")
        else:
            logger.info(f"Title not found for ISBN: {isbn}")
            book_details = {'Title': 'book not found', 'ISBN': isbn}
    else:
        logger.info(f"Book not found for ISBN: {isbn}")
        book_details = {'Title': 'book not found', 'ISBN': isbn}

    return book_details

def extract_book_details(soup):
    try:
        book_details = {}
        
        # Extract Title
        title_tag = soup.find('h1', class_='MuiTypography-root MuiTypography-h1 mui-style-1ngtbwk')
        title = title_tag.text.strip() if title_tag else None
        book_details['Title'] = title
        
        # Extract Author
        author_tag = soup.find('a', href=lambda x: x and 'author=' in x)
        author = author_tag.text.strip() if author_tag else None
        book_details['Author'] = author
    
        # Extract Book Type and Pages
        details = soup.find_all('div', class_='MuiButtonBase-root MuiTab-root MuiTab-labelIcon MuiTab-textColorInherit mui-style-ax6ycu')
        book_type = details[0].text.strip() if len(details) > 0 else None
        pages = details[1].text.strip() if len(details) > 1 else None
        book_details['Book Type'] = book_type
        book_details['Pages'] = pages
        
        # Extract Prices
        original_price_tag = soup.find('p', class_='MuiTypography-root MuiTypography-body1 mui-style-vrqid8')
        original_price = original_price_tag.text.strip().replace('RRP', '').replace('$', '').strip() if original_price_tag else None
        discounted_price_tag = soup.find('p', class_='MuiTypography-root MuiTypography-body1 BuyBox_sale-price__PWbkg mui-style-tgrox')
        discounted_price = discounted_price_tag.text.strip().replace('$', '').strip() if discounted_price_tag else None
        book_details['Original Price'] = original_price
        book_details['Discounted Price'] = discounted_price
        
        # Check for additional details if the main details are not found
        additional_details_tag = soup.find('div', class_='MuiBox-root mui-style-h3npb')
        if additional_details_tag:
            product_details = additional_details_tag.find_all('p', class_='MuiTypography-root MuiTypography-body1 mui-style-tgrox')
            for detail in product_details:
                label_tag = detail.find('span', class_='MuiTypography-root MuiTypography-body1 detail-label mui-style-tgrox')
                if label_tag:
                    label = label_tag.text.strip().replace(':', '')
                    value = detail.text.replace(label_tag.text, '').strip()
                    if label == 'ISBN':
                        book_details['ISBN'] = value
                    elif label == 'ISBN-10':
                        book_details['ISBN-10'] = value
                    elif label == 'Published':
                        book_details['Published Date'] = value
                    elif label == 'Publisher':
                        book_details['Publisher'] = value
                    elif label == 'Number of Pages':
                        book_details['Pages'] = value
                        
    except Exception as e:
        logger.error(f"Error extracting book details: {e}")
        logger.error(traceback.format_exc())  # Log traceback
        return None
    
    return book_details

def export_to_excel(book_details, filename):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Book Details"
        headers = [
            'Title of the Book',
            'Author/s',
            'Book type',
            'Original Price (RRP)',
            'Discounted price',
            'ISBN',
            'ISBN-10',
            'Published Date',
            'Publisher',
            'No. of Pages'
        ]
        sheet.append(headers)
    
    row = [
        book_details.get('Title', ''),
        book_details.get('Author', ''),
        book_details.get('Book Type', ''),
        book_details.get('Original Price', ''),
        book_details.get('Discounted Price', ''),
        book_details.get('ISBN', ''),
        book_details.get('ISBN-10', ''),
        book_details.get('Published Date', ''),
        book_details.get('Publisher', ''),
        book_details.get('Pages', '')
    ]
    sheet.append(row)
    workbook.save(filename)
    logger.info(f"Appended book details to Excel: {filename}")

def main():
    # Read ISBNs from CSV
    csv_filename = 'isbn_input_list.csv'  # Replace with the path to your CSV file
    isbn_list = read_isbn_list_from_csv(csv_filename)
    
    for isbn in isbn_list:
        url = f"https://www.booktopia.com.au/search?keywords={isbn}&productType=917504"
        book_details = scrape_book_details(url, isbn)
        if book_details:
            export_to_excel(book_details, 'z_isbn_book_details.xlsx')
        else:
            export_to_excel({'Title': 'book not found', 'ISBN': isbn}, 'isbn_book_details.xlsx')

if __name__ == '__main__':
    main()
